from __future__ import annotations

import os
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from models import Circunscripcion, EleccionCongreso2023, Partido, ResultadoPartido


class ExcelStructureError(Exception):
    pass


class ElectionDataLoader:
    def __init__(self, excel_path: str) -> None:
        self.excel_path = excel_path

    def load_election(self) -> Tuple[EleccionCongreso2023, List[str]]:
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"No se encontro el archivo Excel en la ruta: {self.excel_path}")

        workbook = load_workbook(self.excel_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if len(rows) < 4:
            raise ExcelStructureError("El Excel no contiene suficientes filas.")

        election = EleccionCongreso2023(
            nombre="Elecciones generales al Congreso 2023",
            archivo_origen=self.excel_path,
            metadatos_columnas={"origen": sheet.title},
        )
        mensajes = [f"CONFIRMACION: Se selecciono la hoja '{sheet.title}'."]
        datos = self._parse_wide_format(rows)
        if not datos:
            raise ExcelStructureError("No se pudieron extraer filas de resultados.")

        for row in datos:
            circ = self._get_or_create_circunscripcion(election, row)
            partido = self._build_partido(row)
            mensajes.append(election.registrar_partido(partido))
            mensajes.append(
                circ.agregar_resultado(
                    ResultadoPartido(
                        partido=partido,
                        votos=int(row["votos"]),
                        escanos_oficiales=int(row.get("escanos_oficiales_partido", 0)),
                    )
                )
            )
        return election, mensajes

    def _parse_wide_format(self, rows: List[Tuple[object, ...]]) -> List[Dict[str, object]]:
        header_index = 2
        party_name_row = list(rows[header_index - 2])
        party_sigla_row = list(rows[header_index - 1])
        headers = [self._as_text(v) for v in rows[header_index]]
        first_party = self._find_first_party_col(party_name_row, party_sigla_row)

        idx_codigo = self._find_header(headers, "codigo de provincia")
        idx_nombre = self._find_header(headers, "nombre de provincia")
        idx_ccaa = self._find_header(headers, "comunidad autonoma")
        idx_poblacion = self._find_header(headers, "poblacion")
        idx_cera = self._find_header(headers, "censo cera")
        idx_votantes_cera = self._find_header(headers, "total votantes cera")
        idx_votantes = self._find_header(headers, "total votantes")
        idx_blanco = self._find_header(headers, "votos en blanco")
        idx_nulos = self._find_header(headers, "votos nulos")
        idx_escanos = self._find_header(headers, "diputados")

        if idx_codigo < 0 or idx_nombre < 0 or idx_escanos < 0:
            raise ExcelStructureError("No se localizaron columnas obligatorias del Excel.")

        salida: List[Dict[str, object]] = []
        for row in rows[header_index + 1 :]:
            codigo = self._normalize_numeric_code(self._value(row, idx_codigo))
            nombre = self._as_text(self._value(row, idx_nombre))
            if not codigo or not nombre:
                continue

            base = {
                "circunscripcion_codigo": codigo,
                "circunscripcion_nombre": nombre,
                "comunidad_autonoma": self._as_text(self._value(row, idx_ccaa)),
                "poblacion": self._to_optional_int(self._value(row, idx_poblacion)),
                "censo_cera": self._to_optional_int(self._value(row, idx_cera)),
                "total_votantes_cera": self._to_optional_int(self._value(row, idx_votantes_cera)),
                "total_votantes": self._to_optional_int(self._value(row, idx_votantes)),
                "votos_blanco": self._to_optional_int(self._value(row, idx_blanco)),
                "votos_nulos": self._to_optional_int(self._value(row, idx_nulos)),
                "escanos_circunscripcion": self._to_int(self._value(row, idx_escanos)),
            }

            for col in range(first_party, len(headers)):
                party_name = self._as_text(party_name_row[col] if col < len(party_name_row) else "")
                party_sigla = self._as_text(party_sigla_row[col] if col < len(party_sigla_row) else "")
                if not party_name and not party_sigla:
                    continue
                votos = self._to_int(self._value(row, col))
                if votos <= 0:
                    continue
                salida.append(
                    {
                        **base,
                        "partido_codigo": party_sigla or party_name,
                        "partido_nombre": party_name or party_sigla,
                        "partido_sigla": party_sigla,
                        "votos": votos,
                        "escanos_oficiales_partido": 0,
                    }
                )
        return salida

    def _get_or_create_circunscripcion(self, election: EleccionCongreso2023, row: Dict[str, object]) -> Circunscripcion:
        codigo = str(row["circunscripcion_codigo"])
        if codigo in election.circunscripciones:
            return election.circunscripciones[codigo]
        circ = Circunscripcion(
            codigo=codigo,
            nombre=str(row["circunscripcion_nombre"]),
            provincia=str(row["circunscripcion_nombre"]),
            comunidad_autonoma=str(row.get("comunidad_autonoma") or "Sin comunidad"),
            escanos_oficiales_totales=int(row.get("escanos_circunscripcion") or 0),
            poblacion=self._to_optional_int(row.get("poblacion")),
            censo_cera=self._to_optional_int(row.get("censo_cera")),
            total_votantes_cera=self._to_optional_int(row.get("total_votantes_cera")),
            total_votantes=self._to_optional_int(row.get("total_votantes")),
            votos_blanco_oficiales=self._to_optional_int(row.get("votos_blanco")),
            votos_nulos_oficiales=self._to_optional_int(row.get("votos_nulos")),
        )
        election.registrar_circunscripcion(circ)
        return circ

    def _build_partido(self, row: Dict[str, object]) -> Partido:
        return Partido(
            codigo=str(row.get("partido_codigo") or "").strip(),
            nombre=str(row.get("partido_nombre") or "").strip(),
            sigla=str(row.get("partido_sigla") or "").strip(),
        )

    def _find_first_party_col(self, row1, row2):
        for i in range(max(len(row1), len(row2))):
            if self._as_text(row1[i] if i < len(row1) else "") or self._as_text(row2[i] if i < len(row2) else ""):
                return i
        return max(len(row1), len(row2))

    def _find_header(self, headers, name):
        target = self._normalize(name)
        for i, header in enumerate(headers):
            if self._normalize(header) == target:
                return i
        return -1

    def _value(self, row, idx):
        if idx < 0 or idx >= len(row):
            return None
        return row[idx]

    def _to_int(self, value):
        if value is None:
            return 0
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(round(value))
        text = str(value).strip().replace(".", "").replace(",", "")
        match = re.search(r"-?\d+", text)
        return int(match.group(0)) if match else 0

    def _to_optional_int(self, value) -> Optional[int]:
        return None if value is None or str(value).strip() == "" else self._to_int(value)

    def _as_text(self, value):
        return "" if value is None else str(value).strip()

    def _normalize(self, value):
        return self._as_text(value).lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")

    def _normalize_numeric_code(self, value):
        text = self._as_text(value)
        digits = "".join(c for c in text if c.isdigit())
        return digits.zfill(2) if digits else text